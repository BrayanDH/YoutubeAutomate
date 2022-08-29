
import shutil
import openpyxl
from pytube import YouTube
from pytube.cli import on_progress
from hurry.filesize import size
import os
import re


# you only need specific your final quality, 720p or 1080p, and the line to you need start from excel archive, finally the archive path
line = 2
video_quality = "720p"

# your path here
path = "E:/Trabajo/Trailers"
min_video_size = 28.00
extension = ".mp4"


# regex to get the name of the file
def change_name(path, name, extension):
    splitted_name = re.sub(r'[^a-zA-Z0-9 _]', r'', name)
    spaces_replace = '_'.join(splitted_name.split())
    final_name = str(spaces_replace) + str(extension)
    os.rename(path, final_name)
    return final_name


# move file to the path
def move_file(file_name, path):
    curren_directory = os.getcwd()
    archive_path = curren_directory + "/" + file_name
    shutil.move(archive_path, path)


# open excel file
wb = openpyxl.load_workbook('links.xlsx')
ws = wb.active
link = ""

while link != None:

    # get link from excel and close if there is no link

    link = ws.cell(row=line, column=1).value
    if link == None:
        break

    # get download object and video metadata
    download = YouTube(link, on_progress_callback=on_progress)
    meta_data = {
        'id_file': line,
        'link': link,
        'title': download.title,
        'length': download.length,
    }

    # format metadata times
    lenght = round(meta_data["length"] / 60, 2)
    convert_time = str(lenght)
    duration = convert_time[0:4] + " min"

    # get video stream and file size
    download = download.streams.filter(res=f"{video_quality}").first()
    try:
        file_size = download.filesize
    except AttributeError:
        file_size = 0

    # check if video is smaller than minimum size
    size_final = size(file_size)
    if lenght < min_video_size:
        print(f"\nThis video has {lenght} mins, it will be downloaded..\n")
        print(f"Video name:{meta_data['title']}")
        print(f"ID:{meta_data['id_file']}")
        print(f"Link: {meta_data['link']}")
        print(f"Size file:{size_final}\n")

        # download video and replace name
        try:
            final_path = path + "/" + "video.mp4"
            download.download(output_path=path, filename=f"video.mp4")

            final_name = str(meta_data['id_file']) + \
                "_" + str(meta_data['title'])
            file_to_move = change_name(final_path, final_name, extension)
            move_file(file_to_move, path)

            # add metadata to excel
            data_to_write = [
                meta_data["id_file"],
                meta_data["link"],
                meta_data["title"],
                size_final,
                duration
            ]
            wb2 = openpyxl.load_workbook('output_register.xlsx')
            ws2 = wb2.active
            column = 1
            for data in data_to_write:
                ws2.cell(row=line, column=column).value = data
                column += 1
            wb2.save("output_register.xlsx")

        except AttributeError:
            print("There was a problem with the download, I'll go to the following link.")
    else:
        print(
            f"\nThis video has {lenght} minutes, it will not download\n")
    line += 1
